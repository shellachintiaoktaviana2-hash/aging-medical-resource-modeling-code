# -*- coding: utf-8 -*-
"""
老龄化熵权法结果 + AHP老龄化质量模型结果合成代码

处理逻辑：
1. 读取两份Excel中“综合得分排名”工作表；
2. 按“地区”匹配两个省级得分；
3. 分别对两个得分进行极差标准化；
4. 对标准化后的两个得分取等权平均，得到最终老龄化综合得分；
5. 按最终得分降序排序并导出Excel。

需要安装：
pip install pandas openpyxl
"""

import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import DataBarRule, ColorScaleRule

# ===================== 1. 修改为你的本地文件路径 =====================
entropy_file = r"老龄化拓展模型_熵权法结果(2).xlsx"
ahp_file = r"AHP老龄化质量模型结果(3).xlsx"

# 输出路径：默认保存在熵权法文件所在文件夹
out_dir = os.path.dirname(os.path.abspath(entropy_file)) or os.getcwd()
out_file = os.path.join(out_dir, "老龄化熵权法_AHP合成最终得分结果表.xlsx")

# ===================== 2. 读取数据 =====================
df_entropy = pd.read_excel(entropy_file, sheet_name="综合得分排名")
df_ahp = pd.read_excel(ahp_file, sheet_name="综合得分排名")

# 自动定位得分列，避免列名略有差异时出错
def find_score_col(df, keywords):
    for col in df.columns:
        name = str(col)
        if all(k in name for k in keywords):
            return col
    raise ValueError(f"未找到包含关键词 {keywords} 的得分列，当前列名为：{list(df.columns)}")

entropy_score_col = find_score_col(df_entropy, ["老龄化", "得分"])
ahp_score_col = find_score_col(df_ahp, ["AHP", "得分"])

# 保留关键列
df_entropy = df_entropy[["地区", "年份", "排名", entropy_score_col]].copy()
df_entropy.columns = ["地区", "年份", "熵权法文件排名", "熵权法原始得分"]

df_ahp = df_ahp[["地区", "排名", ahp_score_col]].copy()
df_ahp.columns = ["地区", "AHP文件排名", "AHP原始得分"]

# ===================== 3. 按地区合并 =====================
df = pd.merge(df_entropy, df_ahp, on="地区", how="inner")

# 检查是否有未匹配地区
missing_entropy = set(df_entropy["地区"]) - set(df["地区"])
missing_ahp = set(df_ahp["地区"]) - set(df["地区"])
if missing_entropy or missing_ahp:
    print("警告：存在未匹配地区。")
    print("熵权法文件中未匹配：", missing_entropy)
    print("AHP文件中未匹配：", missing_ahp)

# ===================== 4. 极差标准化 =====================
# 公式：X'_ij = (X_ij - min(X_j)) / (max(X_j) - min(X_j))
def minmax_standardize(series):
    min_v = series.min()
    max_v = series.max()
    if max_v == min_v:
        return pd.Series([0.0] * len(series), index=series.index)
    return (series - min_v) / (max_v - min_v)

df["熵权法标准化"] = minmax_standardize(df["熵权法原始得分"])
df["AHP标准化"] = minmax_standardize(df["AHP原始得分"])

# ===================== 5. 等权平均合成最终得分 =====================
# 公式：S_i = 0.5 * E'_i + 0.5 * A'_i = (E'_i + A'_i) / 2
df["最终老龄化综合得分"] = (df["熵权法标准化"] + df["AHP标准化"]) / 2

# 得分解释，可根据论文需要自行调整阈值
def score_label(x):
    if x >= 0.75:
        return "高老龄化压力"
    elif x >= 0.50:
        return "中高老龄化压力"
    elif x >= 0.25:
        return "中低老龄化压力"
    else:
        return "低老龄化压力"

df["结果解释"] = df["最终老龄化综合得分"].apply(score_label)

# 排名
df = df.sort_values("最终老龄化综合得分", ascending=False).reset_index(drop=True)
df.insert(0, "最终排名", range(1, len(df) + 1))

# ===================== 6. 生成方法参数表 =====================
params = pd.DataFrame({
    "参数": ["熵权法最小值", "熵权法最大值", "AHP最小值", "AHP最大值", "样本地区数", "熵权法合成权重", "AHP合成权重"],
    "数值": [
        df["熵权法原始得分"].min(),
        df["熵权法原始得分"].max(),
        df["AHP原始得分"].min(),
        df["AHP原始得分"].max(),
        len(df),
        0.5,
        0.5,
    ]
})

method = pd.DataFrame({
    "步骤": [1, 2, 3, 4],
    "处理内容": ["数据匹配", "极差标准化", "等权平均合成", "排序"],
    "公式": [
        "按地区名称一一匹配",
        "X'=(X-min(X))/(max(X)-min(X))",
        "S_i=(E'_i+A'_i)/2",
        "Rank_i=rank(S_i)"
    ],
    "说明": [
        "两个结果表均以省级地区为评价对象。",
        "消除两个模型得分量纲和取值范围差异。",
        "熵权法结果与AHP结果权重均设为0.5。",
        "最终得分越高，综合老龄化压力越强。"
    ]
})

# ===================== 7. 导出Excel =====================
with pd.ExcelWriter(out_file, engine="openpyxl") as writer:
    df.to_excel(writer, sheet_name="合成结果", index=False)
    df[["地区", "年份", "熵权法文件排名", "熵权法原始得分", "AHP文件排名", "AHP原始得分"]].to_excel(
        writer, sheet_name="原始合并数据", index=False
    )
    params.to_excel(writer, sheet_name="参数", index=False)
    method.to_excel(writer, sheet_name="方法说明", index=False)

# ===================== 8. 简单美化 =====================
wb = load_workbook(out_file)
header_fill = PatternFill("solid", fgColor="1F4E79")
header_font = Font(color="FFFFFF", bold=True)
thin = Side(style="thin", color="D9E2F3")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for ws in wb.worksheets:
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if isinstance(cell.value, float):
                cell.number_format = "0.0000"
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 28)

# 对合成结果中的最终得分添加数据条与标准化得分彩色刻度
ws = wb["合成结果"]
# 查找列号
headers = [cell.value for cell in ws[1]]
final_col = headers.index("最终老龄化综合得分") + 1
std1_col = headers.index("熵权法标准化") + 1
std2_col = headers.index("AHP标准化") + 1
last_row = ws.max_row
ws.conditional_formatting.add(
    f"{ws.cell(2, final_col).coordinate}:{ws.cell(last_row, final_col).coordinate}",
    DataBarRule(start_type="min", end_type="max", color="5B9BD5", showValue=True)
)
ws.conditional_formatting.add(
    f"{ws.cell(2, std1_col).coordinate}:{ws.cell(last_row, std2_col).coordinate}",
    ColorScaleRule(start_type="min", start_color="FCE4D6", mid_type="percentile", mid_value=50, mid_color="FFF2CC", end_type="max", end_color="D9EAD3")
)

wb.save(out_file)

print("已完成，输出文件：", out_file)
print(df[["最终排名", "地区", "最终老龄化综合得分", "结果解释"]].head(10))
